import openpyxl
import matplotlib.pyplot as plt
import numpy as np
path = 'base_aeropuerto.xlsx'
archivo = openpyxl.load_workbook(path)

# 0.2 Menú ADMIN
def mnforad():
    print('\n- Bienvenido al menú Administrador -')
    print('\t1. Añadir usuario')
    print('\t2. Modificar contraseña de usuario existente')
    print('\t3. Eliminar última cuenta existente')
    print('\t4. Mostrar usuarios registrados')
    print('\t5. Reportes finales')
    print('\t6. Salir del menú Administrador')
    mtvwa = int(input('-> Teclee el número correspondiente: '))
    return int(mtvwa)


# 0.2.1 ADMIN: Añadir usuario
def mnforadop1():
    archivo.active = 1
    husuarios = archivo.active
    print('\nAÑADIR USUARIO')
    usuario_a = input('Escriba el usuario (3-8 caracteres): ')
    while len(usuario_a) > 8 or len(usuario_a) < 3:
        usuario_a = input('Escriba el usuario (3-8 caracteres): ')
    password_a = input('Escriba la contraseña (3-8 caracteres): ')
    while len(password_a) > 8 or len(password_a) < 3:
        password_a = input('Escriba la contraseña (3-8 caracteres): ')
    husuarios.append([usuario_a, password_a])
    archivo.save(path)
    print('- ¡Usuario registrado! -')


# 0.2.2 ADMIN: Modificar contraseña de usuario existente
def mnforadop2():
    archivo.active = 1
    husuarios = archivo.active
    print('\nMODIFICAR CONTRASEÑA')
    usuario_m = input('Escriba el usuario a buscar: ')
    lista_col = []
    primera_col = husuarios['A']
    for c in primera_col:
        lista_col.append(c.value)  # enlista todos los valores de la primera columna
    while usuario_m not in lista_col:
        print(f"Usuario '{usuario_m}' no se encuentra registrado.")
        usuario_m = input('Escriba el usuario a buscar: ')
    if usuario_m in lista_col:
        print(f"Usuario '{usuario_m}' encontrado en el registro.")
        usuario_fila = lista_col.index(usuario_m)  # no. de fila
        print('Contraseña anterior:', husuarios.cell(row=usuario_fila + 1, column=2).value)
        password_m = input('Escriba la nueva contraseña: ')
        husuarios.cell(row=usuario_fila + 1, column=2).value = password_m
        archivo.save(path)
        print('- ¡Contraseña modificada correctamente! -')


# 0.2.3 ADMIN: Eliminar última cuenta registrada
def mnforadop3():
    archivo.active = 1
    husuarios = archivo.active
    print('\nELIMINAR ÚLTIMA CUENTA REGISTRADA')
    fila_maxima = husuarios.max_row
    husuarios.delete_rows(fila_maxima, 1)
    archivo.save(path)
    print('- ¡La última cuenta registrada fue eliminada! -')


# 0.2.4 ADMIN: Ver usuarios registrados
def mnforadop4():
    archivo.active = 1
    husuarios = archivo.active
    print('\n'+'-'*48)
    print('|    Tabla de previsualización de usuarios     |')
    print('-'*48)
    for t in range(1, 2):
        print('\t '+husuarios.cell(row=t, column=1).value, end='\t\t')
        print('\t'+husuarios.cell(row=t, column=2).value)
    print('-'*48)
    for r in range(2, husuarios.max_row+1):
        for c in range(1, 3):
            try:
                print('\t '+husuarios.cell(row=r, column=c).value, end='  \t\t')
            except TypeError:
                pass
            except "NoneType":
                pass
        print()
    print('-'*48)


# 0.2.5 ADMIN: Ver reportes finales
def mnforadop5():
    print(" - Los reportes de sesión y simulación registrados se desplegarán en breve -")
    archivo.active = 2
    hreportes = archivo.active
    ultima_fila = hreportes['2']
    fila_lista = []  # almacena los valores del eje 'x'
    for u in ultima_fila:
        fila_lista.append(u.value)

    etiquetas = ['Sesiones de usuario', 'Sesiones de ADMIN', 'Total simulaciones']  # nombres eje x
    valores = fila_lista

    x = np.arange(len(etiquetas))  # posición de las etiquetas
    width = 0.30  # ancho de las barras
    fig, ax = plt.subplots()
    rects1 = ax.bar(x - width / 16, valores, width)

    ax.set_ylabel('Cantidad')  # leyenda eje 'y'
    ax.set_title('Cantidad de sesiones iniciadas y simulaciones realizadas')
    ax.set_xticks(x)
    ax.set_xticklabels(etiquetas)

    def autolabel(rects):
        for rect in rects:
            height = rect.get_height()
            ax.annotate('{}'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom')

    autolabel(rects1)
    fig.tight_layout()
    plt.show()
    print(" \n- Reporte finalizado -")


# 0.2.6 ADMIN: Salir del modo ADMIN
def mnforadop6():
    archivo.save(path)
    archivo.close()
    print('\nSALIR - Sesión ADMIN cerrada.')


# 0.1 Login GENERAL
def iniciosecure():
    archivo.active = 1
    husuarios = archivo.active
    intento = 0
    es = 'a'

    usuario = input('USUARIO: ')
    password = input('CONTRASEÑA: ')
    lista_usuarios = []
    col_usuarios = husuarios['A']
    for c in col_usuarios:
        lista_usuarios.append(c.value)
    lista_passwords = []
    col_passwords = husuarios['B']
    for d in col_passwords:
        (lista_passwords.append(str(d.value)))
    while intento < 3:
        if usuario in lista_usuarios and password in lista_passwords:
            archivo.active = 2
            reportes_finales = archivo.active
            cont_adm = reportes_finales.cell(row=2, column=2).value
            cont_adm = cont_adm + 1
            reportes_finales.cell(row=2, column=2).value = cont_adm
            archivo.save(path)
            print(f'      ¡Bienvenid@, {usuario}!')
            break
        elif usuario not in lista_usuarios or password not in lista_passwords:
            intento = intento + 1
            print(f'[!] Utilizó {intento} intento(s) de 4 [!]')
            usuario = input('USUARIO: ')
            password = input('CONTRASEÑA: ')
    if intento == 3:
        es = 'b'
        print(f'[!] Utilizó {intento + 1} intento(s) de 4 - Cuenta bloqueada [!]')
    return es

